# -*- coding: utf-8 -*-
"""
SiteDoc AI - Excel診断データ ジェネレーター
scanner.full_scan() の結果 dict から「全チェック項目の生データ」を .xlsx で出力する。

LP（index.html）が本診断の納品物として謳っている
「スコア付き診断データ（Excel）― 全チェック項目の生データ」の実体。
report.py（Word）と対になる納品物で、生成・保存の流れも合わせてある。

使い方（CLI）:
  python excel_report.py example.com     # reports/ に .xlsx を生成
"""
import io
import os

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import scanner

SEV_LABEL = {"c": "重大", "w": "要改善", "s": "良好"}
SEV_FILL = {
    "c": PatternFill("solid", fgColor="F8D7D7"),
    "w": PatternFill("solid", fgColor="FCEFD2"),
    "s": PatternFill("solid", fgColor="D6F0E6"),
}
SEV_FONT_COLOR = {"c": "D13438", "w": "B97A0C", "s": "1E8E6A"}

_JP_FONT = "Yu Gothic"
HEADER_FILL = PatternFill("solid", fgColor="1F2A44")
HEADER_FONT = Font(bold=True, color="FFFFFF", name=_JP_FONT)
BASE_FONT = Font(name=_JP_FONT, size=10.5)
BOLD_FONT = Font(name=_JP_FONT, size=10.5, bold=True)
TITLE_FONT = Font(name=_JP_FONT, size=14, bold=True)
_THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _grade(score):
    if score >= 75:
        return "A", "良好"
    if score >= 50:
        return "B", "要改善"
    return "C", "危険"


def _autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _styled_header_row(ws, row_idx, headers):
    for c, text in enumerate(headers, start=1):
        cell = ws.cell(row_idx, c, text)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center", horizontal="center")
        cell.border = BORDER


def build_workbook(report):
    """scanレポート dict から Workbook を組み立てて返す。"""
    wb = Workbook()
    g, gtext = _grade(report["overall"])

    # ================= 概要シート =================
    ws1 = wb.active
    ws1.title = "概要"
    ws1["A1"] = "SiteDoc AI — Webセキュリティ診断 概要"
    ws1["A1"].font = TITLE_FONT
    ws1.merge_cells("A1:B1")

    meta_rows = [
        ("診断対象ドメイン", report["domain"]),
        ("診断日時", report.get("scannedAt", "")),
        ("総合スコア", f"{report['overall']} / 100点"),
        ("総合評価", f"{g}（{gtext}）"),
        ("検出：重大", report["counts"]["c"]),
        ("検出：要改善", report["counts"]["w"]),
        ("検出：良好", report["counts"]["s"]),
    ]
    r = 3
    for k, v in meta_rows:
        ws1.cell(r, 1, k).font = BOLD_FONT
        ws1.cell(r, 2, v).font = BASE_FONT
        r += 1

    r += 1
    ws1.cell(r, 1, "領域別スコア").font = BOLD_FONT
    r += 1
    _styled_header_row(ws1, r, ["領域", "スコア"])
    r += 1
    for area in report["areaResults"]:
        ws1.cell(r, 1, area["name"]).font = BASE_FONT
        ws1.cell(r, 2, f"{area['score']} / 100").font = BASE_FONT
        ws1.cell(r, 1).border = BORDER
        ws1.cell(r, 2).border = BORDER
        r += 1

    _autosize(ws1, [22, 42])

    # ================= 診断詳細シート（全チェック項目の生データ）=================
    ws2 = wb.create_sheet("診断詳細")
    headers = ["領域", "領域スコア", "深刻度", "項目", "内容"]
    _styled_header_row(ws2, 1, headers)
    ws2.freeze_panes = "A2"

    row_i = 2
    for area in report["areaResults"]:
        for f in area["findings"]:
            ws2.cell(row_i, 1, area["name"]).font = BASE_FONT
            ws2.cell(row_i, 2, area["score"]).font = BASE_FONT

            sev_cell = ws2.cell(row_i, 3, SEV_LABEL.get(f["sev"], f["sev"]))
            sev_cell.font = Font(name=_JP_FONT, bold=True,
                                  color=SEV_FONT_COLOR.get(f["sev"], "000000"))
            sev_cell.fill = SEV_FILL.get(f["sev"], PatternFill())
            sev_cell.alignment = Alignment(horizontal="center")

            ws2.cell(row_i, 4, f["title"]).font = BOLD_FONT

            desc_cell = ws2.cell(row_i, 5, f["desc"])
            desc_cell.font = BASE_FONT
            desc_cell.alignment = Alignment(wrap_text=True, vertical="top")

            for c in range(1, 6):
                ws2.cell(row_i, c).border = BORDER
            row_i += 1

    _autosize(ws2, [16, 10, 10, 32, 76])
    ws2.auto_filter.ref = f"A1:E{row_i - 1}"

    return wb


def generate_bytes(report):
    """xlsx を bytes で返す（Webダウンロード等に使う場合）。"""
    wb = build_workbook(report)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _out_dir():
    d = os.path.join(os.path.dirname(os.path.abspath(__file__)), "reports")
    os.makedirs(d, exist_ok=True)
    return d


def generate_file(domain):
    """スキャンを実行し、reports/ に .xlsx を保存してパスを返す。"""
    report = scanner.full_scan(domain)
    wb = build_workbook(report)
    safe = domain.replace(":", "_").replace("/", "_")
    path = os.path.join(_out_dir(), f"{safe}_診断データ.xlsx")
    wb.save(path)
    return path


if __name__ == "__main__":
    import sys
    domain = sys.argv[1] if len(sys.argv) > 1 else "example.com"
    ok, reason = scanner.is_public_domain(domain)
    if not ok:
        print(f"診断できません（{reason}）：{domain}")
        sys.exit(1)
    p = generate_file(domain)
    print(f"生成しました：{p}")
